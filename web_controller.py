#!/usr/bin/env python3
from flask import Flask, Response,render_template, request, make_response, send_file
from functools import wraps
import json
import os
import sys
import socket
from ParamFilter import FilterObj as Filter

import ManufacturerManager

from io import BytesIO

import model

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.hyperlink import Hyperlink
import pandas as pd


path = os.path.realpath(os.path.dirname(sys.argv[0]))

#------------------------------
app = Flask(__name__)


def __GetSpec(data):
    result = []
    rows = data.split('\n')

    for row in rows:
        try:
            temp_item = { 
                'name': '', 
                'count': 1, 
                }
            name = row.split('\t')[0]
            if name != "":
                temp_item['name'] = name

            count = row.split('\t')[1]
            if count != "":
                temp_item['count'] = count


            result.append(temp_item)
        except:
            continue

    return result

    


@app.route("/")
def index():
    return render_template('index.html')


@app.route("/version", methods=['GET'])
def getversion():
    res = { 
        'version': model.GetVersion()
    }
    return res


# Маршрут для обработки POST запроса и генерации файла
@app.route('/get_manufacturers_info', methods=['GET'])
def get_manufacturers_info():
    info = ManufacturerManager.NameGenerator()
    res = { 
        'res_smd': info.GetSmdResManufacturers(), 
        'cer_cap_smd': info.GetSmdCerCapManufacturers(), 
        'tant_cap_smd': info.GetSmdTantCapManufacturers(), 
        }
    return res


# Маршрут для обработки POST запроса и генерации файла
@app.route('/download_csv', methods=['POST'])
def download():
    # Получение данных с формы
    name = request.form.get('bom_list')
    # name = name.replace("\t", ";")
    name = name.replace("\r\n", "\n")

    columns = len(name.split("\n")[0].split("\t"))
    # Создание текста для файла

    if (columns < 3):
        file_content = f'Name\tQuantity\n{name}'
    else:
        file_content = f'Name\tQuantity\tManufacturer\n{name}'

    # Создание файла в памяти
    file_stream = BytesIO()
    file_stream.write(file_content.encode('utf-16'))
    file_stream.seek(0)

    # Отправка файла пользователю для скачивания
    return send_file(file_stream, as_attachment=True, download_name="BOM.csv", mimetype='text/plain')



# Маршрут для обработки POST запроса и генерации файла
@app.route('/download_excel', methods=['POST'])
def download_excel():
    parser_filter = Filter()
    data = request.get_json()
    bom = data['bom']
    filter = data['cap_filter']['skip_tol']
    spec_list = __GetSpec(bom)


    device_count = int(data['count'])
    tech_reseve = float(data['tech_res'])

    parser_filter.SetSkipingEndurance('R', data['res_filter']['skip_power'])
    parser_filter.SetSkipingTolerance('R', data['res_filter']['skip_tol'])

    parser_filter.SetSkipingTolerance('C', data['cap_filter']['skip_tol'])
    parser_filter.SetSkipingEndurance('C', data['cap_filter']['skip_voltage'])
    parser_filter.SetSkipingVariant('C', data['cap_filter']['skip_dielectric'])

    res_data = {
        '#': [],
        'Исходное наименование': [],
        'Параметры': [],
        'Список на англ.': [],
        'Список на рус.': [],
        'Наимен. произв.': [],
        'Производитель': [],
        'Количество': [], 
    }

    res_list = []
    count = 1
    for item in spec_list:

        model.CorrectionCount(item, device_count, tech_reseve)

        manufacturers_settings = ManufacturerManager.Settings()

        parse_res = model.HandleRowBOM(item, ['elitan', 'chipdip', 'platan', 'promelec', 'dko_electronshik'], manufacturers_settings, parser_filter)

        res_data['#'].append(count)
        res_data['Исходное наименование'].append(item['name'])

        param_str = ""

        for param in parse_res['params']:
            param_str = f'{param_str}{param}\n'
        
        param_str = param_str[:-1]

        res_data['Параметры'].append(param_str)
        res_data['Количество'].append(int(item['count']))
        res_data['Список на англ.'].append(parse_res['en_text_item'])
        res_data['Список на рус.'].append(parse_res['ru_text_item'])
        res_data['Наимен. произв.'].append(parse_res['manufacturer_info']['component_name'])
        res_data['Производитель'].append(parse_res['manufacturer_info']['manufacturer_name'])


        for store in parse_res['ordering']:

            if store['store_name'] not in res_data:  # Проверка на уникальность ключа
                res_data[store['store_name']] = []
            res_data[store['store_name']].append(store['order_link'])

        count = count + 1



    
    # Преобразуем данные в DataFrame
    df = pd.DataFrame(res_data)

    # Создаем Excel файл в памяти
    
    wb = Workbook()
    ws = wb.active

    # Заполняем рабочий лист данными
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    

    # Применяем форматирование
    for cell in ws[1]:  # Заголовки
        cell.font = Font(bold=True)  # Жирный шрифт для заголовков
        cell.alignment = Alignment(wrap_text=True, vertical='top') 

    ws['A'][0].alignment =  Alignment(wrap_text=True, vertical='top', horizontal='right') 

    # Применяем форматирование (перенос текста) и ширину столбцов
    for cell in ws['A'][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for cell in ws['B'][1:]: 
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for cell in ws['C'][1:]: 
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for cell in ws['D'][1:]: 
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for cell in ws['E'][1:]: 
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for cell in ws['F'][1:]: 
        cell.alignment = Alignment(wrap_text=True, vertical='top') 
    for cell in ws['G'][1:]: 
        cell.alignment = Alignment(wrap_text=True, vertical='top') 
    for cell in ws['H'][1:]: 
        cell.alignment = Alignment(wrap_text=True, vertical='top')   


    for cell in ws['I'][1:]:  # Срез, начинающийся со второй ячейки
        cell.style = "Hyperlink"
        cell.hyperlink = cell.value
        cell.value = "ссылка"
        cell.alignment = Alignment(wrap_text=True, vertical='top')  


    for cell in ws['J'][1:]:  # Срез, начинающийся со второй ячейки
        cell.style = "Hyperlink"
        cell.hyperlink = cell.value
        cell.value = "ссылка"
        cell.alignment = Alignment(wrap_text=True, vertical='top')  

    for cell in ws['K'][1:]:  # Срез, начинающийся со второй ячейки
        cell.style = "Hyperlink"
        cell.hyperlink = cell.value
        cell.value = "ссылка"
        cell.alignment = Alignment(wrap_text=True, vertical='top') 

    for cell in ws['L'][1:]:  # Срез, начинающийся со второй ячейки
        cell.style = "Hyperlink"
        cell.hyperlink = cell.value
        cell.value = "ссылка"
        cell.alignment = Alignment(wrap_text=True, vertical='top')   

    for cell in ws['M'][1:]:  # Срез, начинающийся со второй ячейки
        cell.style = "Hyperlink"
        cell.hyperlink = cell.value
        cell.value = "ссылка"
        cell.alignment = Alignment(wrap_text=True, vertical='top')  
                

    # Задаем ширину столбцов
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 35 
    ws.column_dimensions['C'].width = 30 
    ws.column_dimensions['D'].width = 30 
    ws.column_dimensions['E'].width = 30 
    ws.column_dimensions['F'].width = 30 
    ws.column_dimensions['G'].width = 20 
    ws.column_dimensions['H'].width = 7 
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15

    output = BytesIO()
    # Сохраняем рабочую книгу в память
    wb.save(output)
    output.seek(0)  # Перематываем указатель в начало файла

    # Отправляем файл в ответ на запрос
    return send_file(output, as_attachment=True, download_name="ResultTable1.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    

@app.route('/bom_data', methods=['GET', 'POST'])
def handle_bom():
    parser_filter = Filter()
    bom = request.form.get('bom')
    spec_list = __GetSpec(bom)

    device_count = int(request.form.get('count'))
    tech_reseve = float(request.form.get('tech_res'))

    parser_filter.SetSkipingEndurance('R', request.form.get('res_filter[skip_power]') == 'true')
    parser_filter.SetSkipingTolerance('R', request.form.get('res_filter[skip_tol]') == 'true')

    parser_filter.SetSkipingTolerance('C', request.form.get('cap_filter[skip_tol]') == 'true')
    parser_filter.SetSkipingEndurance('C', request.form.get('cap_filter[skip_voltage]') == 'true')
    parser_filter.SetSkipingVariant('C', request.form.get('cap_filter[skip_dielectric]') == 'true')


    res_list = []
    for item in spec_list:

        model.CorrectionCount(item, device_count, tech_reseve)

        manufacturers_settings = ManufacturerManager.Settings()

        parse_res = model.HandleRowBOM(item, ['elitan', 'chipdip', 'platan', 'promelec', 'dko_electronshik'], manufacturers_settings, parser_filter)

        temp_item = { 
            'name': item['name'], 
            'type': parse_res['type'],
            'count': item['count'],
            'params': parse_res['params'],
            'ordering': parse_res['ordering'],
            'ru': parse_res['ru_text_item'],
            'en': parse_res['en_text_item'],
            'elitan': parse_res['elitan_text_item'],
            'manufacturer_info': parse_res['manufacturer_info']
            }
        res_list.append(temp_item)

    return res_list
    



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5003)